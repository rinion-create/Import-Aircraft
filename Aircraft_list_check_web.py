
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
SafetyManager365 — Aircraft Import Validator & Suggester (v0.5.3, DH8x patch)

Key fixes:
- ICAO suggestions seeded from the import token (e.g., DH8C) so DH8x always maps to DHC-8-xxx types.
- Manufacturer canonicalization: DE HAVILLAND -> DE HAVILLAND CANADA (Master uses De Havilland Canada, not Bombardier).
- Treat generic engines (Turboprop/Jet/Piston...) as "missing" -> fallback to first valid engine model from Master.
- Dash-8 family special case: DH8*/DHC-8-xxx considered same family ('8') to retain candidates.
- Cleaned minor issues in earlier snippet (regex alias struct, function signatures, etc.).
"""

import io
import os
import re
import difflib
import platform
import zipfile
from typing import Optional, Tuple, Set, Dict, List, Iterable

import pandas as pd
import streamlit as st

# ---------------------- UI CONFIG --------------------------------------------
st.set_page_config(
    page_title="SM365 Aircraft Import Validator",
    page_icon="🛫",
    layout="wide",
)

# ---------------------- VERBOSE LOGGING --------------------------------------
def get_logger(verbose: bool):
    logs: List[str] = []

    def info(msg: str):
        if verbose:
            logs.append(f"[INFO] {msg}")

    def debug(msg: str):
        if verbose:
            logs.append(f"[DEBUG] {msg}")

    def flush():
        if logs:
            with st.expander("Logs (verbose)", expanded=False):
                st.write("\n".join(logs))

    return info, debug, flush

# --- Configuration -----------------------------------------------------------

HOME = os.path.expanduser("~")

# macOS OneDrive mount root
MAC_ONEDRIVE_ROOTS = [
    os.path.join(HOME, "Library", "CloudStorage", "OneDrive-Comply365"),
    os.path.join(HOME, "Library", "CloudStorage", "OneDrive - Comply365"),
    os.path.join(HOME, "Library", "CloudStorage", "OneDrive"),
]

# Windows OneDrive root
WIN_ONEDRIVE_ROOTS = [
    os.path.join(HOME, "OneDrive - Comply365"),
    os.path.join(HOME, "OneDrive"),
]

PROJECT_FOLDER = os.path.join(
    "ASQS - Project and Account Management - Project and Account Management",
    "04. Tools"
)

# Fixed expected file names
MASTER_FILENAMES = [
    "260101_Aicraft_Master.xlsx",  # keep typo tolerance
    "260101_Aircraft_Master.xlsx",
]

# Regex patterns (case-insensitive) for broader auto-search
MASTER_REGEX_PATTERNS = [
    r"(?i)\b(?:\d{6}_)?aicraft_master\.xlsx$",   # tolerate date prefix + 'Aicraft' typo
    r"(?i)\b(?:\d{6}_)?aircraft_master\.xlsx$",  # tolerate date prefix
    r"(?i)\b.*aircraft.*master.*\.xlsx$",        # broad fallback
]

# Repo-bundled Master fallback (for cloud)
REPO_MASTER_FALLBACK = os.path.join(os.path.dirname(__file__), "260101_Aircraft_Master.xlsx")

MANUFACTURER_COL_CANDIDATES = ["manufacturer", "aircraft manufacturer", "mfr", "oem", "maker"]
TYPE_COL_CANDIDATES         = ["type", "aircraft type", "model", "family", "series", "aircraft model"]
ENGINE_COL_CANDIDATES       = ["engine", "engine type", "engine model", "powerplant", "motor"]
ID_COL_CANDIDATES           = ["id", "aircraft id", "aircraft_id", "uid", "key", "master id", "aircraft code"]

# Optional ICAO/IATA type designator column candidates (case-insensitive)
TYPE_ICAO_COL_CANDIDATES    = [
    "icao", "icao type", "icao code", "icao designator", "icao type designator",
    "type designator", "icao aircraft type", "icao model", "iata"  # tolerate IATA column too
]

# MTOW column candidates (case-insensitive)
MTOW_COL_CANDIDATES = [
    "mtow", "mtow (kg)", "maximum takeoff weight", "maximum take-off weight",
    "max takeoff weight", "max take-off weight", "mtow kg", "mtow (t)",
    "mtow (lbs)", "mtow lbs",
]

# Defaults
DEFAULT_SUGGESTION_COUNT   = 3
DEFAULT_SUGGESTION_CUTOFF  = 0.60
DEFAULT_MASTER_SEARCH_DEPTH = 6

# --- Helpers ----------------------------------------------------------------

def sanitize_path(p: str) -> str:
    """Strip surrounding quotes, expand ~ and env vars, normalize."""
    if not p:
        return p
    p = p.strip()
    if (p.startswith("'") and p.endswith("'")) or (p.startswith('"') and p.endswith('"')):
        p = p[1:-1]
    p = os.path.expanduser(os.path.expandvars(p))
    return os.path.normpath(p)

def normalize_preserve_case(value) -> str:
    """
    Normalize while preserving case and hyphens:
    - Trim whitespace
    - Unify EN/EM dash to '-' and tighten spaces around hyphens
    - Collapse remaining spaces
    - DO NOT lowercase
    """
    if pd.isna(value):
        return ""
    s = str(value).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s*-\s*", "-", s)
    s = re.sub(r"\s+", " ", s)
    return s

# --- Type-specific normalization --------------------------------------------

TYPE_SPACE_FIXES = [
    (r'(?<=\b[A-Za-z])\s+(?=\d)', ''),    # 'EC 135' -> 'EC135'
    (r'(?<=\d)\s+(?=[A-Za-z])', ''),      # '135 T2 +' -> '135T2 +'
]


# DO NOT remove the + sign — keep C680+ distinct from C680
TYPE_VARIANT_FIXES = [
    (r'\bPLUS\b', '+'),  # convert the word "PLUS" to '+'
    # Remove the rule that strips/mangles '+'
]


def normalize_type(value) -> str:
    """Type-specific normalization keeping case/hyphens and compaction."""
    s = normalize_preserve_case(value)
    for pat, repl in TYPE_SPACE_FIXES:
        s = re.sub(pat, repl, s)
    for pat, repl in TYPE_VARIANT_FIXES:
        s = re.sub(pat, repl, s, flags=re.IGNORECASE)
    return s

# --- Dash‑8 ICAO normalizer (hard) ------------------------------------------
def normalize_dash8_icao(t: str) -> str:
    """
    Cleanup for Dash-8 ICAO types (DH8A/DH8B/DH8C/DH8D).
    Removes spaces/thin spaces, uppercases, truncates to canonical 4-char token.
    """
    if t is None or (isinstance(t, float) and pd.isna(t)):
        return ""
    s = str(t)
    s = s.replace("\u2009", "").replace("\u202F", "").replace(" ", "")
    s = s.upper().strip()
    if s.startswith("DH8") and len(s) >= 4:
        return s[:4]
    return s

# --- Variant synonym harmonization ------------------------------------------
TYPE_SYNONYMS: Dict[str, str] = {
    # e.g., "DHC-8-300 DASH 8": "DHC-8-300",  # optional helper for scoring; writeback uses Master displays
}

def harmonize_type_variants(t: str) -> str:
    """Normalize then apply synonym mapping, plus Dash‑8 ICAO normalization."""
    nt = normalize_type(t)
    nt = normalize_dash8_icao(nt)
    canon_key_upper = TYPE_SYNONYMS.get(nt.upper())
    canon_key_case  = TYPE_SYNONYMS.get(nt)
    canon = canon_key_upper if canon_key_upper is not None else canon_key_case
    return normalize_type(canon) if canon else nt

# --- Manufacturer normalization & aliasing ----------------------------------

def clean_dangling_punct(s: str) -> str:
    """Remove stray trailing/leading punctuation and compress spaces."""
    s = re.sub(r"\s*[,.;:]\s*$", "", s)
    s = re.sub(r"^\s*[,.;:]\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

LEGAL_SUFFIXES = [
    r"\bGmbH\b", r"\bAG\b", r"\bInc\.?\b", r"\bIncorporated\b", r"\bLLC\b",
    r"\bLtd\.?\b", r"\bLimited\b", r"\bSAS\b", r"\bSA\b", r"\bBV\b", r"\bNV\b",
    r"\bS\.?p\.?A\.?\b",
]

LOCALE_TOKENS = [
    r"\bDeutschland\b", r"\bFrance\b", r"\bUSA\b", r"\bUK\b", r"\bGroupe\b",
    r"\bCanada\b",
]

def normalize_manufacturer_import(value: str) -> str:
    """Import rows: remove legal/locale suffixes, fix punctuation."""
    s = normalize_preserve_case(value)
    for pat in LEGAL_SUFFIXES:
        s = re.sub(pat, "", s, flags=re.IGNORECASE)
    for pat in LOCALE_TOKENS:
        s = re.sub(pat, "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    s = clean_dangling_punct(s)
    return s

def normalize_manufacturer_master(value: str) -> str:
    """Master rows: keep legal suffixes; fix punctuation only."""
    s = normalize_preserve_case(value)
    s = re.sub(r"\s+", " ", s).strip()
    s = clean_dangling_punct(s)
    return s

MANUFACTURER_ALIASES: Dict[str, str] = {
    # Eurocopter consolidation
    "EUROCOPTER": "EUROCOPTER",
    "EUROCOPTER DEUTSCHLAND": "EUROCOPTER",
    "EUROCOPTER FRANCE": "EUROCOPTER",

    # MD Helicopters
    "MD HELICOPTERS": "MD HELICOPTERS, INC.",
    "MD HELICOPTERS INC": "MD HELICOPTERS, INC.",
    "MD HELICOPTERS, INC": "MD HELICOPTERS, INC.",
    "MD HELICOPTERS,": "MD HELICOPTERS, INC.",

    # Bell Textron variants (explicit)
    "BELL TEXTRON": "BELL",
    "BELL TEXTRON CANADA": "BELL",
    "BELL TEXTRON CANADA LTD": "BELL",
    "BELL TEXTRON INC": "BELL",

    # De Havilland canonicalization (Master uses DE HAVILLAND CANADA)
    "DE HAVILLAND": "DE HAVILLAND CANADA",
    "DE HAVILLAND CANADA": "DE HAVILLAND CANADA",
    "DE HAVILLAND AIRCRAFT OF CANADA": "DE HAVILLAND CANADA",
    "DE HAVILLAND AIRCRAFT OF CANADA LIMITED": "DE HAVILLAND CANADA",
}

MANUFACTURER_REGEX_ALIASES: List[Tuple[re.Pattern, str]] = [
    (re.compile(r'^\s*Bell\s+Helicopter\s+Textron\b.*', re.IGNORECASE), "BELL"),
    (re.compile(r'^\s*Bell\b.*\bTextron\b.*', re.IGNORECASE), "BELL"),
    (re.compile(r'^\s*Westland\s+Helicopter\b.*', re.IGNORECASE), "AGUSTAWESTLAND"),
    (re.compile(r'^\s*Costr\.?\s*Aeronaut\.?G\.?Agusta\b.*', re.IGNORECASE), "AGUSTAWESTLAND"),
    (re.compile(r'\bAgusta\s*Westland\b', re.IGNORECASE), "AGUSTAWESTLAND"),
    (re.compile(r'\bAgustaWestland\b', re.IGNORECASE), "AGUSTAWESTLAND"),
]

def canonicalize_manufacturer(s: str) -> str:
    """Alias manufacturer via explicit dict, then regex fallbacks."""
    key = s.upper()
    mapped = MANUFACTURER_ALIASES.get(key)
    if mapped:
        return mapped
    for rx, target in MANUFACTURER_REGEX_ALIASES:
        if rx.search(s):
            return target
    return s

def man_eq(cand: str, given: str) -> bool:
    """Case-sensitive first, fallback to case-insensitive."""
    return (cand == given) or (cand.upper() == given.upper())

def canonicalize_manufacturer_with_type(man: str, typ: str) -> str:
    """EC135-era canonicalization to EUROCOPTER; else standard canonicalization."""
    try:
        fam = re.search(r'(\d{2,3})', harmonize_type_variants(typ)).group(1)
    except Exception:
        fam = None
    if fam == '135' and str(harmonize_type_variants(typ)).upper().startswith('EC'):
        return "EUROCOPTER"
    return canonicalize_manufacturer(man)

# --- Column pickers ----------------------------------------------------------

def pick_column(df: pd.DataFrame, candidates: Iterable[str]) -> str:
    """Pick best-matching column name (case-insensitive on headers)."""
    df_cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in df_cols_lower:
            return df_cols_lower[cand.lower()]
    all_lower = list(df_cols_lower.keys())
    for cand in candidates:
        m = difflib.get_close_matches(cand.lower(), all_lower, n=1, cutoff=0.8)
        if m:
            return df_cols_lower[m[0]]
    raise ValueError(
        f"Could not auto-detect any of {list(candidates)}. Available: {list(df.columns)}"
    )

def try_pick_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    """Best-effort version of pick_column: returns None instead of raising."""
    try:
        return pick_column(df, candidates)
    except ValueError:
        return None

# --- ICAO/IATA alias & code normalization -----------------------------------

def normalize_code_alias(code: str) -> Tuple[str, Set[str]]:
    """
    Return a canonical key and a set of alias keys for an aircraft type designator.
    - Canonical key removes spaces/hyphens (case preserved).
    - If pattern is letters? + digits (2-3) + letters? (e.g., 'B763', 'A321', 'MD900', 'H900'),
      DO NOT add pure numeric alias when a letter prefix exists (avoid collisions).
    """
    if not code:
        return "", set()

    raw = str(code).strip()
    key = re.sub(r"[\s\-]+", "", raw)

    aliases: Set[str] = {key}
    m = re.match(r"^([A-Za-z]{0,2})(\d{2,3})([A-Za-z]{0,2})$", key)
    if m:
        prefix = m.group(1) or ""
        num    = m.group(2)
        tail   = m.group(3) or ""
        if tail:
            aliases.add(num + tail)  # e.g., '77W'
        elif not prefix:
            aliases.add(num)         # pure numeric alias ONLY if there is no alpha prefix
    return key, aliases

# --- Engine fallback helper ---------------------------------------------------

def pick_first_engine_for(man: str, typ: str, master: Dict) -> Tuple[str, str]:
    """
    Return (engine_normalized, master_id) for the first suitable (Manufacturer, Type, Engine) combo.
    Deterministic: stable sort by engine string.
    """
    base_pool = master["combo_list_valid_code"] if master.get("type_icao_col") else master["combo_list_all"]

    candidates = [c for c in base_pool if man_eq(c[0], man) and c[1] == typ]
    if not candidates:
        candidates = [c for c in base_pool if man_eq(c[0], man) and c[1].upper() == typ.upper()]

    candidates.sort(key=lambda c: (str(c[2])))

    if candidates:
        m, t, e = candidates[0]
        mid = master["lookup"].get((m, t, e), "")
        return e, mid

    return "", ""

# --- OneDrive Master path resolution ----------------------------------------

def master_path_candidates() -> List[str]:
    roots = MAC_ONEDRIVE_ROOTS if platform.system() == "Darwin" else WIN_ONEDRIVE_ROOTS
    candidates: List[str] = []
    for root in roots:
        base = os.path.join(root, PROJECT_FOLDER)
        for fname in MASTER_FILENAMES:
            candidates.append(os.path.join(base, fname))
    return candidates

def list_existing_onedrive_roots() -> List[str]:
    roots = MAC_ONEDRIVE_ROOTS if platform.system() == "Darwin" else WIN_ONEDRIVE_ROOTS
    return [r for r in roots if os.path.exists(r)]

# --- OneDrive output resolution & save helpers -------------------------------

def resolve_onedrive_output_dir(custom_subdir: Optional[str] = None) -> Optional[str]:
    """Resolve a writable OneDrive output directory (local only)."""
    roots = list_existing_onedrive_roots()
    if not roots:
        return None

    base_root = roots[0]
    if custom_subdir:
        sub = sanitize_path(custom_subdir)
        if os.path.isabs(sub) and sub.startswith(base_root):
            out_dir = sub
        else:
            out_dir = os.path.join(base_root, sub)
    else:
        out_dir = os.path.join(base_root, PROJECT_FOLDER)

    os.makedirs(out_dir, exist_ok=True)
    return os.path.normpath(out_dir)

def save_bytes_to_path(buf: io.BytesIO, target_path: str) -> None:
    """Persist a BytesIO to target_path."""
    buf.seek(0)
    with open(target_path, "wb") as f:
        f.write(buf.read())

def original_file_stem(filename: Optional[str]) -> str:
    """Return the original filename stem (without extension)."""
    try:
        if not filename:
            return "import"
        base = os.path.basename(str(filename).strip().strip("'\""))
        stem, _ = os.path.splitext(base)
        return stem if stem else "import"
    except Exception:
        return "import"

# --- OneDrive Master auto-discovery (local only) -----------------------------

def depth_limited_walk(root: str, max_depth: int):
    """Yield (dirpath, dirnames, filenames) like os.walk but stop descending after max_depth."""
    root = os.path.normpath(root)
    root_sep_count = root.count(os.sep)
    for dirpath, dirnames, filenames in os.walk(root):
        depth = os.path.normpath(dirpath).count(os.sep) - root_sep_count
        if depth >= max_depth:
            dirnames[:] = []
        yield dirpath, dirnames, filenames

@st.cache_data(show_spinner=False, ttl=300)
def find_master_in_onedrive_cached(roots: Tuple[str, ...],
                                   patterns: Tuple[str, ...],
                                   max_depth: int,
                                   stop_after_first: bool) -> List[str]:
    """Cached helper for performance — returns list of matching file paths."""
    compiled = [re.compile(p) for p in patterns]
    matches: List[str] = []
    for root in roots:
        if not os.path.exists(root):
            continue
        for dirpath, _, filenames in depth_limited_walk(root, max_depth=max_depth):
            for fn in filenames:
                if not fn.lower().endswith(".xlsx"):
                    continue
                if any(rx.search(fn) for rx in compiled):
                    full = os.path.join(dirpath, fn)
                    matches.append(full)
                    if stop_after_first:
                        return matches
    return matches

def find_master_candidates_via_auto_search(info, debug,
                                           max_depth: int = DEFAULT_MASTER_SEARCH_DEPTH,
                                           stop_after_first: bool = True) -> List[str]:
    roots = tuple(list_existing_onedrive_roots())
    if not roots:
        info("No OneDrive roots available for auto-search (local).")
        return []
    info(f"Auto-searching Master under OneDrive roots: {roots} (max_depth={max_depth}, stop_after_first={stop_after_first})")
    matches = find_master_in_onedrive_cached(roots, tuple(MASTER_REGEX_PATTERNS), max_depth, stop_after_first)
    info(f"Auto-search found {len(matches)} candidate(s).")
    for i, p in enumerate(matches[:5], 1):
        debug(f"[AutoSearch] Candidate {i}: {p}")
    return matches

# --- Case-insensitive close matches -----------------------------------------

def get_ci_close_matches(term: str, candidates: list, n: int, cutoff: float) -> list:
    """Case-insensitive difflib matching; returns original-cased candidates."""
    t = term.lower()
    lower_to_orig: Dict[str, str] = {}
    lows: List[str] = []
    for c in candidates:
        lc = c.lower()
        if lc not in lower_to_orig:
            lower_to_orig[lc] = c
            lows.append(lc)
    picks = difflib.get_close_matches(t, lows, n=n, cutoff=cutoff)
    return [lower_to_orig[p] for p in picks]

# --- Master loading & lookup building ---------------------------------------

def load_master_from_path(master_path: str, info, debug) -> Dict:
    """Load master file and build lookups from a filesystem path."""
    info(f"Master path: {master_path}")
    try:
        master_df = pd.read_excel(master_path, engine="openpyxl")
        info(f"Master rows loaded: {len(master_df)}")
    except Exception as e:
        raise RuntimeError(f"Failed to read master Excel file: {e}")

    return build_master_lookups(master_df, master_path, info, debug)

def load_master_from_buffer(master_buffer, info, debug) -> Dict:
    """Load master file and build lookups from an uploaded file buffer."""
    try:
        master_df = pd.read_excel(master_buffer, engine="openpyxl")
        info(f"Master rows loaded (uploaded): {len(master_df)}")
    except Exception as e:
        raise RuntimeError(f"Failed to read uploaded master Excel file: {e}")

    return build_master_lookups(master_df, "<uploaded master>", info, debug)


def build_master_lookups(master_df: pd.DataFrame, master_label: str, info, debug) -> Dict:
    """
    Clean, corrected Master lookup builder.
    FIXES:
      - ICAO mapping now uses ICAO column as key source (correct for C680 etc.)
      - valid_code_types correctly tracks which TYPEs have ICAO codes
      - icao_to_types maps ICAO → set(types)
      - type_to_icao maps TYPE → ICAO
    """

    # ---- Column detection ---------------------------------------------------
    man_col       = pick_column(master_df, MANUFACTURER_COL_CANDIDATES)
    type_col      = pick_column(master_df, TYPE_COL_CANDIDATES)
    eng_col       = pick_column(master_df, ENGINE_COL_CANDIDATES)
    id_col        = pick_column(master_df, ID_COL_CANDIDATES)
    type_icao_col = try_pick_column(master_df, TYPE_ICAO_COL_CANDIDATES)

    debug(f"Master columns: man={man_col}, type={type_col}, engine={eng_col}, id={id_col}, icao={type_icao_col}")

    # ---- Normalization ------------------------------------------------------
    master_df["__man__"]    = master_df[man_col].apply(lambda v: canonicalize_manufacturer(normalize_manufacturer_master(v)))
    master_df["__type__"]   = master_df[type_col].apply(harmonize_type_variants)
    master_df["__engine__"] = master_df[eng_col].apply(normalize_preserve_case)

    if type_icao_col:
        master_df["__icao__"] = master_df[type_icao_col].apply(lambda v: str(v).strip().upper())
    else:
        master_df["__icao__"] = ""

    def is_valid_icao(v: str) -> bool:
        return bool(v) and v not in ("N/A", "NONE", "NULL")

    # ---- Build EXACT lookup (M,T,E → ID) -----------------------------------
    lookup: Dict[Tuple[str, str, str], str] = {}
    lookup_display: Dict[Tuple[str, str, str], Tuple[str, str, str]] = {}

    for _, r in master_df.iterrows():
        key = (r["__man__"], r["__type__"], r["__engine__"])
        mid = r[id_col]

        if pd.notna(mid) and key not in lookup:
            lookup[key] = mid
            lookup_display[key] = (
                str(r[man_col]),
                str(r[type_col]),
                str(r[eng_col])
            )

    # ---- Case‑insensitive lookup -------------------------------------------
    lookup_ci: Dict[Tuple[str, str, str], str] = {}
    for k, mid in lookup.items():
        k_ci = tuple(str(x).upper() for x in k)
        if k_ci not in lookup_ci:
            lookup_ci[k_ci] = mid

    # ---- Sets of unique fields ---------------------------------------------
    man_set  = sorted(set(master_df["__man__"]))
    type_set = sorted(set(master_df["__type__"]))
    eng_set  = sorted(set(master_df["__engine__"]))

    info(f"Unique counts — Manufacturers: {len(man_set)}, Types: {len(type_set)}, Engines: {len(eng_set)}")
    info(f"Valid (Manufacturer,Type,Engine) combinations: {len(lookup)}")

    # ---- ICAO mappings ------------------------------------------------------
    type_to_icao: Dict[str, str] = {}         # TYPE → ICAO
    icao_to_types: Dict[str, Set[str]] = {}   # ICAO → {TYPE}
    valid_code_types: Set[str] = set()        # TYPEs that have valid ICAO codes

    if type_icao_col:
        for _, r in master_df.iterrows():
            icao = r["__icao__"]
            typ  = r["__type__"]

            if is_valid_icao(icao):
                # TYPE → ICAO mapping
                type_to_icao[typ] = icao

                # ICAO → TYPE mapping
                icao_to_types.setdefault(icao, set()).add(typ)

                # TYPE is ICAO-supported
                valid_code_types.add(typ)

        debug(f"Loaded ICAO→types keys: {len(icao_to_types)}")
        debug(f"TYPEs with valid ICAO: {len(valid_code_types)}")

    # ---- Build combination lists -------------------------------------------
    combo_list_all = list(lookup.keys())

    if type_icao_col:
        combo_list_valid_code = [c for c in combo_list_all if c[1] in valid_code_types]
    else:
        combo_list_valid_code = combo_list_all[:]

    # ---- Done ---------------------------------------------------------------
    debug(f"Combos (all): {len(combo_list_all)}, combos (valid ICAO only): {len(combo_list_valid_code)}")

    return {
        "df": master_df,
        "man_col": man_col,
        "type_col": type_col,
        "eng_col": eng_col,
        "id_col": id_col,
        "type_icao_col": type_icao_col,

        "lookup": lookup,
        "lookup_ci": lookup_ci,
        "lookup_display": lookup_display,

        "man_set": man_set,
        "type_set": type_set,
        "eng_set": eng_set,

        "type_to_icao": type_to_icao,
        "icao_to_types": icao_to_types,
        "valid_code_types": valid_code_types,

        "combo_list_all": combo_list_all,
        "combo_list_valid_code": combo_list_valid_code,
        "combo_ids": lookup,
        "master_path": master_label,
    }


# --- Scoring helpers ---------------------------------------------------------

def score_combos_against(man: str, typ: str, eng: str, combos: list, combo_ids: dict, top_n: int, cutoff: float):
    """Return top matches with ratio >= cutoff: [(ratio, (m,t,e), master_id), ...]"""
    target = f"{man} {typ} {eng}".lower()
    scored = []
    for combo in combos:
        m, t, e = combo
        cand = f"{m} {t} {e}".lower()
        ratio = difflib.SequenceMatcher(None, target, cand).ratio()
        if ratio >= cutoff:
            scored.append((ratio, combo, combo_ids[combo]))
    scored.sort(key=lambda x: x[0], reverse=True)
    return scored[:top_n]

def field_weighted_score(man, typ, eng, cand_m, cand_t, cand_e) -> float:
    """Weight manufacturer and type higher than engine."""
    w_man = 0.40
    w_typ = 0.45
    w_eng = 0.15
    return (
        difflib.SequenceMatcher(None, man.lower(), cand_m.lower()).ratio() * w_man +
        difflib.SequenceMatcher(None, typ.lower(), cand_t.lower()).ratio() * w_typ +
        difflib.SequenceMatcher(None, eng.lower(), cand_e.lower()).ratio() * w_eng
    )

# --- Family & prefix helpers -------------------------------------------------

_DIGIT_FAMILY_RX = re.compile(r'(\d{2,3})')  # variant-level by default

def family_key(t: str) -> str:
    tn = harmonize_type_variants(t)
    m = _DIGIT_FAMILY_RX.search(tn)
    return m.group(1) if m else ''

def numeric_family(t: str) -> str:
    tn = harmonize_type_variants(t)
    # Dash-8 special case — treat any DH8/DHC-8 variant as family '8'
    if tn.upper().startswith(("DH8", "DHC8", "DHC-8")):
        return "8"
    m = _DIGIT_FAMILY_RX.search(tn)
    return m.group(1) if m else ''

def is_variant_type(t: str) -> bool:
    x = str(t).replace("-", "")
    if re.match(r"^[A-Za-z]+\d+", x):
        return True
    if re.search(r"[A-Za-z]+$", x) and not x.isdigit():
        return True
    return False

def type_prefix(t: str) -> str:
    tn = harmonize_type_variants(t)
    x = re.sub(r'[\s\-]+', '', str(tn))
    m = re.match(r'^([A-Za-z]+)', x)
    return m.group(1).upper() if m else ''

# --- ICAO override for FIX C (improved) -------------------------------------

ICAO_FAMILY_UNLOCK_CODES: Set[str] = {"DH8A", "DH8B", "DH8C", "DH8D", "D328"}

def is_icao_family_unlock(import_type: str, suggested_type: str, master: Dict) -> bool:
    """
    Override activates when:
      - import type is one of the ICAO unlock codes, OR
      - suggested type resolves via master ICAO to one of those codes, OR
      - suggested type matches Dash-8/DHC-8 patterns while import is DH8*/D328.
    """
    try:
        imp = str(harmonize_type_variants(import_type)).upper()
        sug_norm = str(harmonize_type_variants(suggested_type))

        if imp in ICAO_FAMILY_UNLOCK_CODES:
            return True

        icao_code = master.get("type_to_icao", {}).get(sug_norm, "")
        if icao_code:
            key, _aliases = normalize_code_alias(icao_code)
            if key.upper() in ICAO_FAMILY_UNLOCK_CODES:
                return True

        sug_up = sug_norm.upper()
        if imp.startswith("DH8") and (sug_up.startswith("DHC8") or sug_up.startswith("DHC-8") or "DASH 8" in sug_up):
            return True
        if imp == "D328" and ("328" in sug_up or "DORN" in sug_up):
            return True

    except Exception:
        pass
    return False

# --- Processing core ---------------------------------------------------------

GENERIC_ENGINE_CLASSES = {"TURBOPROP", "JET", "PISTON", "TURBOJET", "TURBOFAN"}
def is_generic_engine(value: str) -> bool:
    s = normalize_preserve_case(value)
    return s.upper() in GENERIC_ENGINE_CLASSES

def process_import(
    import_df: pd.DataFrame,
    master: Dict,
    suggestion_count: int,
    suggestion_cutoff: float,
    info, debug,
    orig_stem: str,
):
    # Detect import columns
    info("Detecting Manufacturer/Type/Engine columns in import...")
    try:
        imp_man_col  = pick_column(import_df, MANUFACTURER_COL_CANDIDATES)
        imp_type_col = pick_column(import_df, TYPE_COL_CANDIDATES)
        imp_eng_col  = pick_column(import_df, ENGINE_COL_CANDIDATES)
    except ValueError as e:
        raise RuntimeError(f"Error detecting columns in import file: {e}")
    imp_id_col = try_pick_column(import_df, ID_COL_CANDIDATES)

    # Registration column (first in Validation report)
    imp_reg_col = try_pick_column(import_df, [
        "registration", "aircraft registration", "reg", "tail", "tail number"
    ])
    debug(f"Registration column detected: {imp_reg_col}")

    # Normalize to internal helper columns
    df = import_df.copy()
    df["__type__"]   = df[imp_type_col].apply(harmonize_type_variants)
    df["__man__"]    = df.apply(lambda r: canonicalize_manufacturer_with_type(
        normalize_manufacturer_import(r[imp_man_col]),
        r["__type__"]
    ), axis=1)
    df["__engine__"] = df[imp_eng_col].apply(normalize_preserve_case)

    total          = len(df)
    match_count    = 0
    warn_count     = 0

    validation_rows: List[Dict[str, str]] = []
    suggested_import = import_df.copy()

    info(f"Validating {total} rows...")
    progress = st.progress(0)
    for idx, row in df.iterrows():
        man_raw  = import_df.at[idx, imp_man_col]
        type_raw = import_df.at[idx, imp_type_col]
        eng_raw  = import_df.at[idx, imp_eng_col]
        reg_raw  = import_df.at[idx, imp_reg_col] if imp_reg_col else ""

        type_raw_norm = harmonize_type_variants(type_raw)

        # --- ICAO TYPE FALLBACK FOR TYPES LIKE C680 / C25A / C56X / LJ45 ---
        fallback_icao = ""
        m = re.match(r"^[A-Z]{1,4}\d{0,3}[A-Z]?$", type_raw_norm.replace("-", ""))
        if m:
            fallback_icao = type_raw_norm.upper()

        # Store fallback for ICAO alias expansion
        fallback_icao_key, fallback_icao_aliases = normalize_code_alias(fallback_icao)

        man = row["__man__"]
        typ = row["__type__"]
        eng = row["__engine__"]

        debug(f"[Row {idx}] Raw:   MAN='{man_raw}', TYPE='{type_raw}', ENG='{eng_raw}', REG='{reg_raw}'")
        debug(f"[Row {idx}] Norm:  MAN='{man}', TYPE='{typ}', ENG='{eng}'")

        key = (man, typ, eng)
        combo_ok  = key in master["lookup"]
        master_id = master["lookup"].get(key, "")

        # Case-insensitive exact lookup fallback
        if not combo_ok and master.get("lookup_ci"):
            key_ci = (man.upper(), typ.upper(), eng.upper())
            if key_ci in master["lookup_ci"]:
                combo_ok = True
                master_id = master["lookup_ci"][key_ci]
                # Preserve canonical normalized key
                for k, v in master["lookup"].items():
                    if v == master_id:
                        key = k
                        break

        man_ok  = man in master["man_set"]
        type_ok = typ in master["type_set"]
        eng_ok  = eng in master["eng_set"]

        # Field-level suggestions (diagnostic only)
        type_universe_for_suggestions = (
            [t for t in master["type_set"] if t in master["valid_code_types"]]
            if master["type_icao_col"] else master["type_set"]
        )
        man_suggestions  = [] if man_ok  else get_ci_close_matches(man, master["man_set"],  suggestion_count, suggestion_cutoff)
        type_suggestions = [] if type_ok else get_ci_close_matches(typ, type_universe_for_suggestions, suggestion_count, suggestion_cutoff)
        eng_suggestions  = [] if eng_ok  else get_ci_close_matches(eng, master["eng_set"],  suggestion_count, suggestion_cutoff)

        # ---------- ICAO/IATA-aware mapping ----------
        icao_type_suggestions: List[str] = []
        icao_combo_hits: List[Tuple[float, Tuple[str, str, str], str]] = []

        imp_icao_candidates: Set[str] = set()
        imp_icao_key, imp_icao_aliases = normalize_code_alias(type_raw_norm)
        if imp_icao_key:
            imp_icao_candidates = {imp_icao_key} | imp_icao_aliases

        if master["type_icao_col"]:
            code_val = master["type_to_icao"].get(typ, "")
            debug(f"[Row {idx}] ICAO designator for type '{typ}': '{code_val}'")

            alias_candidates: Set[str] = set()

            # add Master-derived ICAO if available
            if code_val:
                code_key, code_aliases = normalize_code_alias(code_val)
                alias_candidates |= ({code_key} | code_aliases)

            # always add import fallback (C680, C25B, C56X, etc.)
            if fallback_icao_key:
                alias_candidates |= ({fallback_icao_key} | fallback_icao_aliases)

            # add import token (e.g., DH8C or C680)
            alias_candidates |= imp_icao_candidates

            # EXPAND: Dash‑8 crosswalk (optional, keeps DH8X<->DH1/2/3/4 working)
            DH8_ICAO_TO_IATA = {"DH8A": "DH1", "DH8B": "DH2", "DH8C": "DH3", "DH8D": "DH4"}
            DH8_IATA_TO_ICAO = {v: k for k, v in DH8_ICAO_TO_IATA.items()}
            if imp_icao_key in DH8_ICAO_TO_IATA: alias_candidates.add(DH8_ICAO_TO_IATA[imp_icao_key])
            if imp_icao_key in DH8_IATA_TO_ICAO: alias_candidates.add(DH8_IATA_TO_ICAO[imp_icao_key])

            debug(f"[Row {idx}] Types from ICAO/IATA map: {sorted(set(icao_type_suggestions))[:5]}")
            
            for a in alias_candidates:
                icao_type_suggestions.extend(sorted(master["icao_to_types"].get(a, set())))

            if icao_type_suggestions:
                for t_candidate in set(icao_type_suggestions):
                    for (m, t, e) in master["combo_list_all"]:
                        if t == t_candidate:
                            score = field_weighted_score(man, typ, eng, m, t, e)
                            icao_combo_hits.append((score, (m, t, e), master["lookup"][(m, t, e)]))
                icao_combo_hits.sort(key=lambda x: x[0], reverse=True)

        # ---------- FAMILY DETECTION & LOCKING ----------
        fam_key = numeric_family(type_raw_norm)
        prefix_raw = type_prefix(type_raw_norm)

        family_types = {t for t in master["type_set"] if numeric_family(t) == fam_key}

        if prefix_raw:
            # Only enforce prefix when the candidate TYPE itself has a prefix.
            # Many business-jet master types start with digits (e.g., "680 Citation Sovereign") and would be wrongly filtered out.
            filtered_by_prefix = {t for t in family_types if type_prefix(t) == prefix_raw}
            if filtered_by_prefix:
                family_types = filtered_by_prefix
            else:
                # If nothing matches the prefix, skip prefix filtering entirely (numeric-leading descriptive types).
                debug(
                    f"[Row {idx}] Skipping prefix filter for family '{fam_key}' because master types are digit-leading.")

        if master["type_icao_col"]:
            family_types_with_code = {t for t in family_types if t in master["valid_code_types"]}
            if family_types_with_code:
                family_types = family_types_with_code

        icao_family_num = ''
        if master["type_icao_col"] and icao_type_suggestions:
            icao_family_num = numeric_family(icao_type_suggestions[0])

        # ---------- Fuzzy scoring sets ----------
        filtered_combos: List[Tuple[str, str, str]] = []
        if family_types:
            filtered_combos = [c for c in master["combo_list_valid_code"] if (c[1] in family_types)]
            if not filtered_combos:
                filtered_combos = [c for c in master["combo_list_all"] if (c[1] in family_types)]

        same_manufacturer_combos = [c for c in filtered_combos if man_eq(c[0], man)]
        if same_manufacturer_combos:
            filtered_combos = same_manufacturer_combos

        if icao_family_num and filtered_combos:
            filtered_combos = [c for c in filtered_combos if numeric_family(c[1]) == icao_family_num]

        global_combos = master["combo_list_valid_code"][:]
        if icao_family_num:
            global_combos = [c for c in global_combos if numeric_family(c[1]) == icao_family_num]
        same_manufacturer_global = [c for c in global_combos if man_eq(c[0], man)]
        if same_manufacturer_global:
            global_combos = same_manufacturer_global

        global_scored = [] if combo_ok else score_combos_against(
            man, typ, eng, global_combos, master["combo_ids"], top_n=suggestion_count, cutoff=suggestion_cutoff
        )

        filtered_scored = [] if combo_ok else score_combos_against(
            man, typ, eng, filtered_combos, master["combo_ids"],
            top_n=suggestion_count,
            cutoff=suggestion_cutoff
        )

        # ---------- Original-type-first ----------
        original_type_hits = []
        if (not combo_ok) and (type_raw_norm in family_types):
            preferred_type_combos = [c for c in master["combo_list_valid_code"] if (c[1] == type_raw_norm)]
            if not preferred_type_combos:
                preferred_type_combos = [c for c in master["combo_list_all"] if (c[1] == type_raw_norm)]
            if preferred_type_combos:
                original_type_hits = score_combos_against(
                    man, typ, eng, preferred_type_combos, master["combo_ids"], top_n=suggestion_count, cutoff=suggestion_cutoff
                )

        # ---------- Base-type-first ----------
        base_type_hits = []
        if (not combo_ok) and (not original_type_hits) and family_types:
            base_family_types = [t for t in family_types if not is_variant_type(t)]
            if base_family_types:
                base_type_combos = [c for c in master["combo_list_valid_code"] if (c[1] in base_family_types)]
                if not base_type_combos:
                    base_type_combos = [c for c in master["combo_list_all"] if (c[1] in base_family_types)]
                if base_type_combos:
                    base_type_hits = score_combos_against(
                        man, typ, eng, base_type_combos, master["combo_ids"], top_n=suggestion_count, cutoff=suggestion_cutoff
                    )

        # ---------- Canonical family preference ----------
        canonical_ranked = []
        if (not combo_ok) and filtered_combos and not original_type_hits and not base_type_hits:
            target_ci = f"{man} {typ} {eng}".lower()
            for (m, t, e) in filtered_combos:
                base_score  = difflib.SequenceMatcher(None, target_ci, f"{m} {t} {e}".lower()).ratio()
                shape_score = difflib.SequenceMatcher(None, str(type_raw_norm).lower(), t.lower()).ratio()
                penalty_var = 0.25 if is_variant_type(t) else 0.0
                penalty_family = 0.40 if numeric_family(type_raw_norm) != numeric_family(t) else 0.0
                man_bonus = 0.05 if man_eq(m, man) else 0.0
                final_score = (base_score * 0.4) + (shape_score * 0.6) + man_bonus - penalty_var - penalty_family
                canonical_ranked.append((final_score, (m, t, e), master["lookup"][(m, t, e)]))
            canonical_ranked.sort(reverse=True, key=lambda x: x[0])

        # ---------- Choose suggestion ----------
        suggestion_source = ""
        sug_combo: Tuple[str, str, str] = key  # default to normalized input
        sug_mid   = ""

        if combo_ok:
            match_count += 1
            sug_combo = key
            sug_mid   = master_id
            suggestion_source = "exact"
        else:
            warn_count += 1
            if original_type_hits:
                ratio, sug_combo, sug_mid = original_type_hits[0]
                suggestion_source = f"original_type_preferred(ratio={ratio:.3f})"
            elif base_type_hits:
                ratio, sug_combo, sug_mid = base_type_hits[0]
                suggestion_source = f"base_family_type_preferred(ratio={ratio:.3f})"
            elif icao_combo_hits:
                score, candidate_combo, candidate_mid = icao_combo_hits[0]
                # accept if manufacturers equal AFTER canonicalization
                cand_man_canon = canonicalize_manufacturer(candidate_combo[0])
                imp_man_canon  = canonicalize_manufacturer(man)
                if man_eq(cand_man_canon, imp_man_canon):
                    sug_combo = candidate_combo
                    sug_mid   = candidate_mid
                    suggestion_source = f"icao_weighted(score={score:.3f})"
                elif canonical_ranked:
                    ratio, sug_combo, sug_mid = canonical_ranked[0]
                    suggestion_source = f"canonical_family_choice(score={ratio:.3f})"
                elif filtered_scored:
                    ratio, sug_combo, sug_mid = filtered_scored[0]
                    suggestion_source = f"family_restricted_combo(ratio={ratio:.3f})"
                elif global_scored:
                    ratio, sug_combo, sug_mid = global_scored[0]
                    suggestion_source = f"closest_combo_valid_code_only(ratio={ratio:.3f})"
                else:
                    suggestion_source = "none"
            elif canonical_ranked:
                ratio, sug_combo, sug_mid = canonical_ranked[0]
                suggestion_source = f"canonical_family_choice(score={ratio:.3f})"
            elif filtered_scored:
                ratio, sug_combo, sug_mid = filtered_scored[0]
                suggestion_source = f"family_restricted_combo(ratio={ratio:.3f})"
            elif global_scored:
                ratio, sug_combo, sug_mid = global_scored[0]
                suggestion_source = f"closest_combo_valid_code_only(ratio={ratio:.3f})"
            else:
                suggestion_source = "none"

        # ---------- FIX C: HARD NUMERIC FAMILY LOCK (RELAXED VIA ICAO OVERRIDE) ----------
        if suggestion_source != "none":
            if numeric_family(sug_combo[1]) != numeric_family(type_raw_norm):
                if is_icao_family_unlock(type_raw_norm, sug_combo[1], master):
                    debug(f"[Row {idx}] FIX C relaxed via ICAO unlock for import='{type_raw_norm}' -> suggested='{sug_combo[1]}'")
                else:
                    suggestion_source = "none(family_mismatch)"
                    sug_combo = key
                    sug_mid   = ""

        debug(f"[Row {idx}] → Suggestion source: {suggestion_source}")
        debug(f"[Row {idx}] → Suggested combo (normalized): {sug_combo} | MasterID={sug_mid if sug_mid else '—'}")

        # ---------- Engine-aware writeback ----------
        sug_man_norm, sug_typ_norm, sug_eng_norm = sug_combo
        engine_missing_in_import = (
            pd.isna(eng_raw) or normalize_preserve_case(eng_raw) == "" or is_generic_engine(eng_raw)
        )

        if engine_missing_in_import and suggestion_source != "none":
            if not sug_eng_norm or normalize_preserve_case(sug_eng_norm) == "":
                fallback_eng_norm, fallback_mid = pick_first_engine_for(sug_man_norm, sug_typ_norm, master)
                if fallback_eng_norm:
                    sug_eng_norm = fallback_eng_norm
                    if not sug_mid:
                        sug_mid = fallback_mid
                    sug_combo = (sug_man_norm, sug_typ_norm, sug_eng_norm)

        # ---------- Write display strings ----------
        if suggestion_source == "none":
            sug_man_disp, sug_typ_disp, sug_eng_disp = str(man_raw), str(type_raw), str(eng_raw)
            sug_mid = ""
        else:
            disp = master["lookup_display"].get(sug_combo)
            if disp:
                sug_man_disp, sug_typ_disp, sug_eng_disp = disp
            else:
                sug_man_disp, sug_typ_disp, sug_eng_disp = sug_man_norm, sug_typ_norm, sug_eng_norm

        suggested_import.at[idx, imp_man_col] = sug_man_disp
        suggested_import.at[idx, imp_type_col] = sug_typ_disp
        suggested_import.at[idx, imp_eng_col]  = sug_eng_disp
        if imp_id_col is not None:
            suggested_import.at[idx, imp_id_col] = sug_mid

        # ---------- Validation row (Registration FIRST) ----------
        validation_rows.append({
            "AircraftRegistration": reg_raw,
            "RowIndex": idx,
            "Manufacturer": man_raw,
            "Type": type_raw,
            "Engine": eng_raw,
            "Manufacturer_norm": man,
            "Type_norm": typ,
            "Engine_norm": eng,
            "FieldManufacturerExists": man_ok,
            "FieldTypeExists": type_ok,
            "FieldEngineExists": eng_ok,
            "ComboMatch": combo_ok,
            "MasterID": master_id,
            "SuggestionsManufacturer": "; ".join(man_suggestions),
            "SuggestionsType": "; ".join(type_suggestions),
            "SuggestionsTypeFromICAO": "; ".join(icao_type_suggestions),
            "SuggestionsEngine": "; ".join(eng_suggestions),
            "SuggestionsCombo_CodeRestricted": " | ".join([f"{m} {t} {e}" for _, (m,t,e), _ in filtered_scored]),
            "SuggestionsCombo_Global": " | ".join([f"{m} {t} {e}" for _, (m,t,e), _ in global_scored]),
            "SuggestionsComboFromICAO": " | ".join([f"{m} {t} {e}" for _, (m,t,e), _ in icao_combo_hits]),
            "ChosenSuggestionSource": suggestion_source,
            "ChosenManufacturer": sug_man_disp,
            "ChosenType": sug_typ_disp,
            "ChosenEngine": sug_eng_disp,
            "ChosenMasterID": sug_mid,
        })

        progress.progress(min((idx + 1) / max(total, 1), 1.0))

    # --- Build outputs in-memory --------------------------------------------

    # Validation report (single sheet only)
    report_df = pd.DataFrame(validation_rows)

    validation_bytes = io.BytesIO()
    with pd.ExcelWriter(validation_bytes, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="Validation", index=False)
    validation_bytes.seek(0)

    # Hide selected columns + RED highlight for unmatched (ChosenMasterID empty/NaN)
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill

        wb = load_workbook(validation_bytes)
        ws = wb["Validation"]
        header_row = [cell.value for cell in ws[1]]

        cols_to_hide = [
            "Manufacturer_norm",
            "Type_norm",
            "Engine_norm",
            "SuggestionsType",
            "SuggestionsTypeFromICAO",
            "SuggestionsEngine",
            "SuggestionsCombo_CodeRestricted",
            "SuggestionsComboFromICAO",
        ]

        name_to_idx = {str(v): i+1 for i, v in enumerate(header_row) if v is not None}
        for name in cols_to_hide:
            if name in name_to_idx:
                col_letter = get_column_letter(name_to_idx[name])
                ws.column_dimensions[col_letter].hidden = True

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        if "ChosenMasterID" in name_to_idx:
            id_col_idx = name_to_idx["ChosenMasterID"]
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=id_col_idx).value
                if val in ("", None, "—") or (isinstance(val, str) and val.strip().lower() in ("nan", "none", "null")):
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill

        out_validation = io.BytesIO()
        wb.save(out_validation)
        out_validation.seek(0)
    except Exception:
        out_validation = validation_bytes

    # --- Suggested Import (keeps exact original header order & structure)
    # Uppercase Manufacturer ONLY for the export (SuggestedImport)
    export_suggested = suggested_import.copy()
    export_suggested[imp_man_col] = export_suggested[imp_man_col].apply(
        lambda v: (str(v).upper() if pd.notna(v) else v)
    )

    # --- MTOW digits-only sanitization applied ONLY to export ---------------
    mtow_col = try_pick_column(export_suggested, MTOW_COL_CANDIDATES)
    if mtow_col:
        def digits_only(v):
            if pd.isna(v):
                return v
            s = re.sub(r'[^0-9]', '', str(v))
            return s if s != "" else ""
        export_suggested[mtow_col] = export_suggested[mtow_col].apply(digits_only)

    out_suggested = io.BytesIO()
    with pd.ExcelWriter(out_suggested, engine="openpyxl") as writer:
        export_suggested.to_excel(writer, sheet_name="SuggestedImport", index=False)
    out_suggested.seek(0)

    # RED highlight in SuggestedImport based on exact Master DISPLAY values only
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb2 = load_workbook(out_suggested)
        ws2 = wb2["SuggestedImport"]

        headers = [cell.value for cell in ws2[1]]

        def find_col_index(name: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if isinstance(h, str) and h.strip().lower() == str(name).strip().lower():
                    return i + 1
            return None

        man_col_idx = find_col_index("Manufacturer")      or find_col_index(imp_man_col)
        typ_col_idx = find_col_index("Aircraft Type")     or find_col_index(imp_type_col)
        eng_col_idx = find_col_index("Engine Type")       or find_col_index(imp_eng_col)

        valid_master_display_keys_upper: Set[Tuple[str, str, str]] = set(
            (str(m).upper(), str(t), str(e))
            for (m, t, e) in master["lookup_display"].values()
        )

        if man_col_idx and typ_col_idx and eng_col_idx:
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for row_idx in range(2, ws2.max_row + 1):
                man_disp = ws2.cell(row=row_idx, column=man_col_idx).value
                typ_disp = ws2.cell(row=row_idx, column=typ_col_idx).value
                eng_disp = ws2.cell(row=row_idx, column=eng_col_idx).value

                disp_tuple_upper_man = (
                    (str(man_disp).upper() if man_disp is not None else ""),
                    (str(typ_disp) if typ_disp is not None else ""),
                    (str(eng_disp) if eng_disp is not None else "")
                )

                if disp_tuple_upper_man not in valid_master_display_keys_upper:
                    for col_idx in range(1, ws2.max_column + 1):
                        ws2.cell(row=row_idx, column=col_idx).fill = red_fill

        styled_suggested = io.BytesIO()
        wb2.save(styled_suggested)
        styled_suggested.seek(0)
        out_suggested = styled_suggested
    except Exception:
        pass

    sugg_name = f"{orig_stem}_import_ready.xlsx"
    report_name = f"{orig_stem}_validation_report.xlsx"

    metrics = {
        "total": total,
        "matches": match_count,
        "warnings": warn_count,
        "report_name": report_name,
        "sugg_name": sugg_name,
    }

    return out_validation, out_suggested, metrics, report_df, suggested_import

# ---------------------- Cloud detection --------------------------------------

def running_in_cloud() -> bool:
    """Lightweight detection for Streamlit Cloud / missing OneDrive roots."""
    runtime_hint = os.getenv("STREAMLIT_RUNTIME", "")
    server_enabled = os.getenv("STREAMLIT_SERVER_ENABLED") == "1"
    no_onedrive_roots = len(list_existing_onedrive_roots()) == 0
    return server_enabled or runtime_hint.startswith("streamlit") or no_onedrive_roots

# ---------------------- UI LAYOUT -------------------------------------------

st.title("🛫 SafetyManager365")
st.subheader("Aircraft Import Validator & Suggester")
st.caption("Ruben Inion v0.5.3 (DH8x patch)")

with st.sidebar:
    st.header("Settings")
    verbose = st.toggle("Verbose logging", value=True)
    suggestion_count = st.slider("Suggestion count", min_value=1, max_value=10, value=DEFAULT_SUGGESTION_COUNT, step=1)
    suggestion_cutoff = st.slider("Similarity cutoff", min_value=0.30, max_value=0.90, value=DEFAULT_SUGGESTION_CUTOFF, step=0.05)

    st.divider()
    st.subheader("Master source")
    master_upload = st.file_uploader("Upload Master Excel (.xlsx) (overrides everything)", type=["xlsx"])
    auto_search_master = st.checkbox("Local: Auto-search Master across OneDrive (fallback)", value=True)
    max_search_depth = st.slider(
        "Auto-search depth", min_value=2, max_value=12,
        value=DEFAULT_MASTER_SEARCH_DEPTH, step=1,
        help="Limits recursion when scanning your OneDrive to find the Master file."
    )

    st.divider()
    st.subheader("Output")
    save_to_onedrive = st.checkbox(
        "Local: Save to OneDrive",
        value=False,
        help="Save output files to your OneDrive (local only)"
    )
    onedrive_subfolder = st.text_input(
        "OneDrive subfolder (optional)",
        value=PROJECT_FOLDER,
        help="Relative to your OneDrive root. Leave as default to use the project tools folder."
    )

info, debug, flush_logs = get_logger(verbose)

col1, col2 = st.columns(2)
with col1:
    import_upload = st.file_uploader("Upload Import Excel (.xlsx)", type=["xlsx"])
with col2:
    st.markdown("**Master source status**")
    master_source_placeholder = st.empty()

run_btn = st.button("▶️ Validate & Suggest", type="primary", disabled=(import_upload is None))

# Resolve master
master_dict: Optional[Dict] = None
master_status_lines: List[str] = []

cloud = running_in_cloud()
master_status_lines.append("Cloud mode detected." if cloud else "Local mode detected.")

if master_upload is not None:
    try:
        master_dict = load_master_from_buffer(master_upload, info, debug)
        master_status_lines.append("Using **uploaded Master**.")
    except Exception as e:
        st.error(f"Failed to load uploaded Master: {e}")
else:
    if cloud:
        if os.path.exists(REPO_MASTER_FALLBACK):
            try:
                master_dict = load_master_from_path(REPO_MASTER_FALLBACK, info, debug)
                master_status_lines.append(f"Using repo-bundled Master: `{REPO_MASTER_FALLBACK}`")
            except Exception as e:
                st.error(f"Failed to read repo-bundled Master: {e}")
                master_status_lines.append("Please upload the Master in the sidebar.")
        else:
            master_status_lines.append("Repo-bundled Master not found. Please upload the Master in the sidebar.")
    else:
        candidates = master_path_candidates()
        existing = [p for p in candidates if os.path.exists(p)]
        if existing:
            master_status_lines.append("Fixed-path Master (local):")
            master_status_lines.append(f"- {existing[0]}")
            try:
                master_dict = load_master_from_path(existing[0], info, debug)
                master_status_lines.append(f"Master auto-detected (fixed path): `{existing[0]}`")
            except Exception as e:
                st.error(f"Failed to read detected Master: {e}")
        else:
            master_status_lines.append("No fixed-path OneDrive Master found.")
            if auto_search_master:
                matches = find_master_candidates_via_auto_search(info, debug, max_depth=DEFAULT_MASTER_SEARCH_DEPTH, stop_after_first=True)
                if matches:
                    try:
                        master_dict = load_master_from_path(matches[0], info, debug)
                        master_status_lines.append(f"Master found via **OneDrive auto-search**: `{matches[0]}`")
                    except Exception as e:
                        st.error(f"Failed to read auto-searched Master: {e}")
                else:
                    master_status_lines.append("Auto-search did not find a Master file. You can upload the Master in the sidebar.")
            else:
                master_status_lines.append("Auto-search disabled. You can upload the Master in the sidebar.")

master_source_placeholder.markdown("\n\n".join(master_status_lines))

# Run processing
if run_btn and import_upload is not None:
    try:
        import_df = pd.read_excel(import_upload, engine="openpyxl")
        info(f"Import rows loaded: {len(import_df)}")

        if master_dict is None:
            st.error("Master dataset is not available. Please upload a Master file.")
        else:
            import_stem = original_file_stem(getattr(import_upload, "name", None))

            out_validation, out_suggested, metrics, report_df, suggested_import = process_import(
                import_df=import_df,
                master=master_dict,
                suggestion_count=suggestion_count,
                suggestion_cutoff=suggestion_cutoff,
                info=info,
                debug=debug,
                orig_stem=import_stem,
            )

            # Optional OneDrive Save (local only)
            if save_to_onedrive and not cloud:
                output_dir = resolve_onedrive_output_dir(onedrive_subfolder)
                if output_dir is None:
                    st.warning("OneDrive root not found locally. Files were not saved.")
                    info("OneDrive save skipped: no OneDrive root detected.")
                else:
                    sugg_path = os.path.join(output_dir, metrics["sugg_name"])
                    report_path = os.path.join(output_dir, metrics["report_name"])
                    try:
                        save_bytes_to_path(out_suggested, sugg_path)
                        save_bytes_to_path(out_validation, report_path)
                        out_suggested.seek(0)
                        out_validation.seek(0)
                        st.success(f"Saved to OneDrive:\n- {sugg_path}\n- {report_path}")
                        info(f"Saved Suggested Import to {sugg_path}")
                        info(f"Saved Validation Report to {report_path}")
                    except Exception as e:
                        st.error(f"Failed to save to OneDrive: {e}")
                        info(f"OneDrive save failed: {e}")

            # Metrics
            st.success("Processing complete.")

            # Count rows with missing ChosenMasterID
            unable_to_import = report_df["ChosenMasterID"].isna().sum() + \
                               (report_df["ChosenMasterID"].astype(str).str.strip().isin(
                                   ["", "nan", "None", "null", "—"])).sum()

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Rows validated", metrics["total"])
            m2.metric("Exact matches", metrics["matches"])
            m3.metric("Warnings", metrics["warnings"])
            m4.metric("Unable to Import", unable_to_import)

            # ---------------------- Preview (Modified Suggested Import) ----------------------
            st.subheader("Import_ready preview (modified MTOW)")
            try:
                out_suggested.seek(0)
                preview_df = pd.read_excel(
                    out_suggested,
                    engine="openpyxl",
                    sheet_name="SuggestedImport"
                )
                st.dataframe(preview_df.head(50), use_container_width=True)
            except Exception as e:
                st.warning(f"Unable to render the Import_ready preview: {e}")

            # ---------------------- Downloads --------------------------------------------
            st.subheader("Downloads")
            cdl1, cdl2, cdl3 = st.columns(3)

            with cdl1:
                st.download_button(
                    label=f"⬇️ Download Validation Report ({metrics['report_name']})",
                    data=out_validation,
                    file_name=metrics["report_name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            with cdl2:
                st.download_button(
                    label=f"⬇️ Download Suggested Import ({metrics['sugg_name']})",
                    data=out_suggested,
                    file_name=metrics["sugg_name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            with cdl3:
                out_validation.seek(0)
                out_suggested.seek(0)

                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr(metrics["report_name"], out_validation.read())
                    out_validation.seek(0)
                    zf.writestr(metrics["sugg_name"], out_suggested.read())
                    out_suggested.seek(0)

                zip_buf.seek(0)

                st.download_button(
                    label="⬇️ Download Both (ZIP)",
                    data=zip_buf,
                    file_name=f"{original_file_stem(getattr(import_upload, 'name', None))}_outputs.zip",
                    mime="application/zip",
                    help="Downloads both the validation report and the suggested import together.",
                )

    except Exception as e:
        st.error(f"Processing failed: {e}")

    flush_logs()
